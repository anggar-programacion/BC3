#Desarrollado por Angel Garrido Buron en julio de 2024
#Porgrama que abre archivos bc3 y extrae la informacion en formato de excel

import os
import tkinter as tk
from tkinter import filedialog
import openpyxl
import textwrap
from datetime import datetime


# Función para verificar si un valor es numérico
def es_numero(valor, decimales=None):
    try:
        numero = float(valor)
        if decimales is not None:
            numero_redondeado = round(numero, decimales)
            return numero_redondeado
        return numero
    except ValueError:
        return valor

def es_numero_b(valor, decimales=None):
    try:
        numero = float(valor)
        if decimales is not None:
            numero_redondeado = round(numero, decimales)
            return numero_redondeado
        return numero
    except ValueError:
        return 1.0


def es_fecha(valor):
    try:
        if len(valor) != 6:
            raise ValueError("Longitud de cadena incorrecta. Se esperan 6 caracteres.")
            print(valor)
        
        dia = int(valor[:2])
        mes = int(valor[2:4])
        anno = 2000 + int(valor[4:6])  # Suponiendo que los años están en formato YY

        fecha = datetime(anno, mes, dia)
        fecha_formateada = fecha.strftime("%d/%m/%Y")  # Formatear la fecha como DD/MM/YYYY
        return fecha_formateada

    except ValueError as e:
        print(f"Error al convertir a fecha: {e}")
        return None

def leer_primer_registro_V(archivo):
    codificaciones = ['utf-8', 'latin-1', 'iso-8859-1']
    etiquetas = ["REGISTRO", "PROPIEDAD_ARCHIVO", "VERSION_FORMATO", "PROGRAMA_EMISION", "CABECERA",
                 "JUEGO_CARACTERES", "COMENTARIO", "TIPO_INFORMACION",
                 "NUMERO_CERTIFICACION", "FECHA_CERTIFICACION", "URL_BASE"]
    informacion = {}

    for codificacion in codificaciones:
        try:
            with open(archivo, 'rb') as f:
                primera_linea = f.readline().decode(codificacion).strip()

                campos = primera_linea.split('|')

                for etiqueta, campo in zip(etiquetas, campos):
                    if etiqueta == "TIPO_INFORMACION":
                        tipo_informacion = campo.strip()
                        mapeo = {
                            '1': 'BASE DE DATOS',
                            '2': 'PRESUPUESTO',
                            '3': 'CERTIFICACIÓN (A ORIGEN)',
                            '4': 'ACTUALIZACIÓN DE BASE DE DATOS'
                        }
                        tipo_texto = mapeo.get(tipo_informacion, "TIPO NO ESPECIFICADO")
                        informacion[etiqueta] = tipo_texto
                    else:
                        informacion[etiqueta] = campo.strip()
                
                return informacion
        except UnicodeDecodeError:
            continue

    print("No se pudo decodificar el archivo con ninguna codificación compatible.")
    return informacion

def contar_registros_por_tipo(archivo):
    tipos_validos = "VKCDYRTPLQJWGEOMNIABF"
    contador_registros = {tipo: 0 for tipo in tipos_validos}
    total_registros = 0

    try:
        with open(archivo, 'rb') as f:
            for linea in f:
                if len(linea) >= 2 and linea[0:1] == b'~':
                    tipo_registro = chr(linea[1])
                    if tipo_registro in tipos_validos:
                        contador_registros[tipo_registro] += 1
                        total_registros += 1
    except FileNotFoundError:
        print(f"Error: Archivo {archivo} no encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo {archivo}: {e}")
    
    return contador_registros, total_registros

def leer_registros_C_sin_clasificar(archivo, textos_descriptivos):
    equivalencias_tipo = {
        "0": "SIN CLASIFICAR",
        "1": "MANO DE OBRA",
        "2": "MAQUINARIA Y MEDIOS AUXILIARES",
        "3": "MATERIALES",
        "4": "COMPONENTES ADICIONALES DE RESIDUO",
        "5": "CLASIFICACIÓN DE RESIDUO"
    }

    datos = []
    descripcion_dict = {texto[0]: texto[1] for texto in textos_descriptivos}

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo_codigo = campos[6].strip()
                tipo = equivalencias_tipo.get(tipo_codigo, "DESCONOCIDO")

                precio=es_numero(precio)
                fecha=es_fecha(fecha)

                #SI EL TIPO DE CODIGO ES 0 ES QUE ES SIN CLASIFICAR
                if tipo_codigo != "0":
                    continue

                if codigo.endswith('##') or codigo.endswith('#'):
                    continue

                resumen_lines = textwrap.fill(resumen, width=50)
                texto_descriptivo = descripcion_dict.get(codigo, "NO DISPONIBLE")

                datos.append([codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo])

    return datos
    #return codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo


def leer_registros_C_mano_obra(archivo, textos_descriptivos):
    equivalencias_tipo = {
        "0": "SIN CLASIFICAR",
        "1": "MANO DE OBRA",
        "2": "MAQUINARIA Y MEDIOS AUXILIARES",
        "3": "MATERIALES",
        "4": "COMPONENTES ADICIONALES DE RESIDUO",
        "5": "CLASIFICACIÓN DE RESIDUO"
    }

    datos = []
    descripcion_dict = {texto[0]: texto[1] for texto in textos_descriptivos}

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo_codigo = campos[6].strip()
                tipo = equivalencias_tipo.get(tipo_codigo, "DESCONOCIDO")

                precio=es_numero(precio)
                fecha=es_fecha(fecha)

                #SI EL TIPO DE CODIGO ES 1 ES QUE ES MANO DE OBRA
                if tipo_codigo != "1":
                    continue

                if codigo.endswith('##') or codigo.endswith('#'):
                    continue

                resumen_lines = textwrap.fill(resumen, width=50)
                texto_descriptivo = descripcion_dict.get(codigo, "NO DISPONIBLE")

                datos.append([codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo])

    return datos
    #return codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo

def leer_registros_C_materiales(archivo, textos_descriptivos):
    equivalencias_tipo = {
        "0": "SIN CLASIFICAR",
        "1": "MANO DE OBRA",
        "2": "MAQUINARIA Y MEDIOS AUXILIARES",
        "3": "MATERIALES",
        "4": "COMPONENTES ADICIONALES DE RESIDUO",
        "5": "CLASIFICACIÓN DE RESIDUO"
    }

    datos = []
    descripcion_dict = {texto[0]: texto[1] for texto in textos_descriptivos}

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo_codigo = campos[6].strip()
                tipo = equivalencias_tipo.get(tipo_codigo, "DESCONOCIDO")
                
                precio=es_numero(precio)
                fecha=es_fecha(fecha)

                if tipo_codigo != "3":
                    continue

                resumen_lines = textwrap.fill(resumen, width=50)
                texto_descriptivo = descripcion_dict.get(codigo, "NO DISPONIBLE")

                datos.append([codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo])

    return datos

def leer_registros_C_Maquinaria(archivo, textos_descriptivos):
    equivalencias_tipo = {
        "0": "SIN CLASIFICAR",
        "1": "MANO DE OBRA",
        "2": "MAQUINARIA Y MEDIOS AUXILIARES",
        "3": "MATERIALES",
        "4": "COMPONENTES ADICIONALES DE RESIDUO",
        "5": "CLASIFICACIÓN DE RESIDUO"
    }

    datos = []
    descripcion_dict = {texto[0]: texto[1] for texto in textos_descriptivos}

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo_codigo = campos[6].strip()
                tipo = equivalencias_tipo.get(tipo_codigo, "DESCONOCIDO")
                
                precio=es_numero(precio)
                fecha=es_fecha(fecha)

                if tipo_codigo != "2":
                    continue

                resumen_lines = textwrap.fill(resumen, width=50)
                texto_descriptivo = descripcion_dict.get(codigo, "NO DISPONIBLE")

                datos.append([codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo])

    return datos


def leer_registros_C_residuos(archivo, textos_descriptivos):
    equivalencias_tipo = {
        "0": "SIN CLASIFICAR",
        "1": "MANO DE OBRA",
        "2": "MAQUINARIA Y MEDIOS AUXILIARES",
        "3": "MATERIALES",
        "4": "COMPONENTES ADICIONALES DE RESIDUO",
        "5": "CLASIFICACIÓN DE RESIDUO"
    }

    datos = []
    descripcion_dict = {texto[0]: texto[1] for texto in textos_descriptivos}

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo_codigo = campos[6].strip()
                tipo = equivalencias_tipo.get(tipo_codigo, "DESCONOCIDO")
                
                precio=es_numero(precio)
                fecha=es_fecha(fecha)

                if tipo_codigo != "4":
                    continue

                resumen_lines = textwrap.fill(resumen, width=50)
                texto_descriptivo = descripcion_dict.get(codigo, "NO DISPONIBLE")

                datos.append([codigo, unidad, resumen_lines, precio, fecha, texto_descriptivo])

    return datos

def leer_registros_T(archivo_bc3):
    registros_T = []

    with open(archivo_bc3, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~T'):
                try:
                    linea_decodificada = linea.decode('utf-8').strip()
                except UnicodeDecodeError:
                    linea_decodificada = linea.decode('iso-8859-1').strip()
                campos = linea_decodificada.split('|')

                if len(campos) >= 3:
                    codigo_concepto = campos[1].strip()
                    texto_descriptivo = '|'.join(campos[2:]).strip()
                    registros_T.append([codigo_concepto, texto_descriptivo])

    return registros_T

def extraer_lineas_capitulo(archivo):
    lineas_capitulos = []
    nombre_medicion = None

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('||')
                if len(campos) >= 2:
                    primer_campo = campos[0].split('|')
                    if len(primer_campo) >= 2:
                        if primer_campo[1].endswith('##'):
                            nombre_medicion = primer_campo[1].rstrip('#')
                        elif primer_campo[1].endswith('#'):
                            codigo_capitulo = primer_campo[1].rstrip('#')
                            titulo_y_extras = campos[1].split('#')[0]
                            partes_titulo = titulo_y_extras.split('|')

                            titulo_capitulo = partes_titulo[0].strip() if len(partes_titulo) > 0 else ""
                            precio = partes_titulo[1].strip() if len(partes_titulo) > 1 else "0"
                            precio = es_numero(precio)
                            fecha = partes_titulo[2].strip() if len(partes_titulo) > 2 else ""
                            fecha = es_fecha(fecha)
                            tipo = partes_titulo[3].strip() if len(partes_titulo) > 3 else ""
                            tipo = es_numero(tipo)

                            lineas_capitulos.append([codigo_capitulo, titulo_capitulo, precio, fecha, tipo])

    #print (lineas_capitulos)
    return lineas_capitulos, nombre_medicion

def extraer_registros_K(archivo):
    registros_K = {}
    encabezados_primer_bloque = ["DN", "DD", "DS", "DR", "DI", "DP", "DC", "DM", "DIVISA"]
    encabezados_segundo_bloque = ["CI", "GG", "BI", "BAJA", "IVA"]
    encabezados_tercer_bloque = ["DRC", "DC2", "", "DFS", "DRS", "", "DUO", "DI2", "DES", "DN2", "DD2", "DS2", "DSP", "DEC", "DIVISA2"]

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~K'):
                registro_K = linea.decode('iso-8859-1').strip()
                bloques = registro_K.split('|')

                # Inicializar bloques vacíos si faltan
                primer_bloque = bloques[1].split('\\') if len(bloques) > 1 else []
                segundo_bloque = bloques[2].split('\\') if len(bloques) > 2 else []
                tercer_bloque = bloques[3].split('\\') if len(bloques) > 3 else []
                n_bloque = bloques[4].strip() if len(bloques) > 4 else ""

                valores = {}

                for i, valor in enumerate(primer_bloque):
                    if i < len(encabezados_primer_bloque):
                        valores[encabezados_primer_bloque[i]] = valor.strip()

                for i, valor in enumerate(segundo_bloque):
                    if i < len(encabezados_segundo_bloque):
                        valores[encabezados_segundo_bloque[i]] = valor.strip()

                for i, valor in enumerate(tercer_bloque):
                    if i < len(encabezados_tercer_bloque):
                        valores[encabezados_tercer_bloque[i]] = valor.strip()

                valores["n"] = n_bloque

                registros_K[registro_K] = valores

    return registros_K


def leer_registros_C(archivo):
    datos_C = {}
    equivalencias_tipo = {}  # Define este diccionario según tus necesidades

    with open(archivo, 'rb') as f:
        for linea in f:
            if linea.startswith(b'~C'):
                campos = linea.decode('iso-8859-1').strip().split('|')
                codigo = campos[1].strip()
                unidad = campos[2].strip()
                resumen = campos[3].strip()
                precio = campos[4].strip()
                fecha = campos[5].strip()
                tipo = equivalencias_tipo.get(campos[6].strip(), "Desconocido")
                datos_C[codigo] = {
                    "unidad": unidad,
                    "resumen": resumen,
                    "precio": precio,
                    "fecha": fecha,
                    "tipo": tipo
                }

    return datos_C

def leer_registros_M(archivo, datos_C):
    lineas_medicion = []
    total_medicion = 0
    sumas_parciales_por_codigo = {}
    sumas_parciales_por_capitulo = {}
    lineas_por_codigo = {}
    lineas_por_capitulo = {}
    total_codigos=0
    total_capitulos=0


    with open(archivo, 'r', encoding='iso-8859-1') as f:
        for index, linea in enumerate(f):
            if linea.startswith('~M'):
                linea_sin_m = linea[2:].strip()
                campos = linea_sin_m.split('|')
                campo1 = campos[1].strip()
                partes = campo1.split('\\')
                cod_capitulo = partes[0].split("#")[0]
                codigo = partes[1]
                posicion = campos[2].strip().split('\\')
                posi1 = posicion[0]
                posi2 = posicion[1]
                medicion_parcial = es_numero(campos[3].strip(), 2)
                resto_campos = campos[4].strip()

                # Comprobar si existe el campo etiqueta
                if len(campos) > 5:
                    etiqueta = campos[5].strip()
                else:
                    etiqueta = ""

                subcampos = resto_campos.strip().split('\\')
                tipo = subcampos[0].strip()
                subcampos = subcampos[1:]

                num_elementos_por_bloque = 6
                for i in range(0, len(subcampos), num_elementos_por_bloque):
                    bloque = subcampos[i:i + num_elementos_por_bloque]
                    if len(bloque) == num_elementos_por_bloque:
                        comentario = bloque[0].strip()
                        unidades = es_numero_b(bloque[1].strip(), 2)
                        longitud = es_numero_b(bloque[2].strip(), 2)
                        latitud = es_numero_b(bloque[3].strip(), 2)
                        altura = es_numero_b(bloque[4].strip(), 2)
                        clave = bloque[5].strip()
                        medicion_calculada = es_numero(unidades * longitud * latitud * altura, 2)

                        # Buscar en los datos de ~C
                        datos_c = datos_C.get(codigo, {"unidad": "", "resumen": "", "precio": ""})
                        unidad = datos_c["unidad"]
                        resumen = datos_c["resumen"]
                        precio = es_numero(datos_c["precio"])
                        # Calculamos el precio de la linea
                        parcial_linea = es_numero(medicion_calculada * precio, 2)
                        total_medicion += parcial_linea

                        # Acumular la suma por código
                        if codigo not in sumas_parciales_por_codigo:
                            sumas_parciales_por_codigo[codigo] = 0
                            lineas_por_codigo[codigo] = []
                        sumas_parciales_por_codigo[codigo] += parcial_linea
                        lineas_por_codigo[codigo].append(len(lineas_medicion))
                        

                        # Acumular la suma por capítulo
                        if cod_capitulo not in sumas_parciales_por_capitulo:
                            sumas_parciales_por_capitulo[cod_capitulo] = 0
                            lineas_por_capitulo[cod_capitulo] = []
                        sumas_parciales_por_capitulo[cod_capitulo] += parcial_linea
                        lineas_por_capitulo[cod_capitulo].append(len(lineas_medicion))


                        # Crear la línea de medición con la estructura correcta
                        lineas_medicion.append([
                            cod_capitulo, posi1, posi2, medicion_parcial, codigo, resumen, precio, unidad, comentario,
                            unidades, longitud, latitud, altura, medicion_calculada, parcial_linea, "", "", clave])
    
    # Añadir la suma parcial por código a la última línea en la posición correcta
    for codigo, indices in lineas_por_codigo.items():
        if indices:  # Verifica que haya al menos una línea
            ultima_linea_indice = indices[-1]
            parcial_codigo = es_numero(sumas_parciales_por_codigo[codigo], 2)
            lineas_medicion[ultima_linea_indice][-3] = parcial_codigo
#            total_codigo += parcial_codigo

    # Añadir la suma parcial por capítulo a la última línea en la posición correcta
    for capitulo, indices in lineas_por_capitulo.items():
        if indices:  # Verifica que haya al menos una línea
            ultima_linea_indice = indices[-1]
            parcial_capitulo = es_numero(sumas_parciales_por_capitulo[capitulo], 2)
            lineas_medicion[ultima_linea_indice][-2] = parcial_capitulo
#            total_capitulos += parcial_capitulo

    return lineas_medicion, total_medicion,parcial_codigo,parcial_capitulo


def escribir_en_excel(nombre_archivo, contador_por_tipo, total_registros, informacion_v, sin_clasificar_datos, mano_obra_datos, materiales_datos, maquinaria_datos, residuos_datos,textos_descriptivos_datos, capitulos_datos, nombre_medicion, registros_K, lineas_medicion,total_medicion,parcial_codigo,parcial_capitulo):
    workbook = openpyxl.Workbook()


    sheet1 = workbook.create_sheet("NOMBRE MEDICIÓN") #CREA UNA HOJA EN EL LIBRO LLAMADA 
    sheet1.append(["NOMBRE MEDICIÓN"]) #AÑADE UNA LINEA CON EL NOMBRE Y LA MEDICION COMO CABECERAS
    sheet1.append([nombre_medicion]) #AÑADE LAS LINEAS CON ESA INFO

    sheet2 = workbook.create_sheet("INFORMACIÓN V") 
    for etiqueta, valor in informacion_v.items():
        sheet2.append([etiqueta, valor])

    
    sheet3 = workbook.create_sheet("CONTEO POR TIPO")
    sheet3.append(["TIPO", "CANTIDAD"])
    for tipo, cantidad in contador_por_tipo.items():
        sheet3.append([tipo, es_numero(cantidad)])
    sheet3.append(["TOTAL", total_registros])

    

    sheet4 = workbook.create_sheet("SIN CLASIFICAR")
    sheet4.append(["CÓDIGO", "UNIDAD", "RESUMEN", "PRECIO", "FECHA", "TEXTO DESCRIPTIVO"])
    for dato in sin_clasificar_datos:
        sheet4.append(dato)


    sheet5 = workbook.create_sheet("MANO DE OBRA")
    sheet5.append(["CÓDIGO", "UNIDAD", "RESUMEN", "PRECIO", "FECHA", "TEXTO DESCRIPTIVO"])
    for dato in mano_obra_datos:
        sheet5.append(dato)

    sheet6 = workbook.create_sheet("MATERIALES")
    sheet6.append(["CÓDIGO", "UNIDAD", "RESUMEN", "PRECIO", "FECHA", "TEXTO DESCRIPTIVO"])
    for dato in materiales_datos:
        sheet6.append(dato)

    sheet7 = workbook.create_sheet("MAQUINARIA")
    sheet7.append(["CÓDIGO", "UNIDAD", "RESUMEN", "PRECIO", "FECHA", "TEXTO DESCRIPTIVO"])
    for dato in maquinaria_datos:
        sheet7.append(dato)

    sheet8 = workbook.create_sheet("RESIDUOS")
    sheet8.append(["CÓDIGO", "UNIDAD", "RESUMEN", "PRECIO", "FECHA", "TEXTO DESCRIPTIVO"])
    for dato in residuos_datos:
        sheet8.append(dato)


    sheet9 = workbook.create_sheet("TEXTOS DESCRIPTIVOS")
    sheet9.append(["CÓDIGO CONCEPTO", "TEXTO DESCRIPTIVO"])
    for texto in textos_descriptivos_datos:
        sheet9.append(texto)

    sheet12 = workbook.create_sheet("MEDICION")
    sheet12.append(["CAPITULO", "POS_1", "POS_2", "MEDICION PARCIAL", "CÓDIGO", "TEXTO RESUMEN", "PRECIO PARTIDA","UNIDAD", "COMENTARIO LINEA", "UDS", "LARGO", "ANCHO", "ALTO", "MEDICION CALCULADA","PARCIAL LINEA", "PARCIAL CODIGO", "PARCIAL CAPITULO", "CLAVE"])
    
    
    for dato in lineas_medicion:
        sheet12.append(dato)
  
    
    #Añadir la fila del total al final
    sheet12.append(["", "", "", "", "", "", "", "", "", "", "TOTALES","","","",total_medicion, parcial_codigo,parcial_capitulo,""])
    #print (total_medicion)


    sheet10 = workbook.create_sheet("CAPÍTULOS")
    sheet10.append(["CAPÍTULO", "TÍTULO", "SUBTITULO","SUBTITULO2","SUBTITULO3","PRECIO","FECHA","TIPO"])

    total_precio = 0.0
    
    for capitulo in capitulos_datos:
        codigo, titulo, precio, fecha, tipo = capitulo
        if codigo.startswith("-"):
            sheet10.append([codigo, None, None, None, None, None, None, None])
        elif len(codigo) == 1:
            sheet10.append([codigo, titulo, None, None, None,precio, fecha, tipo,])
        elif len(codigo) == 2:
            sheet10.append([codigo, None, titulo, None,None, precio, fecha, tipo])
        elif len(codigo) == 3:
            sheet10.append([codigo, None, None, titulo, None, precio, fecha, tipo])
        elif len(codigo) == 4:
            sheet10.append([codigo, None, None, None, titulo, precio, fecha, tipo])
        elif len(codigo) == 5:
            sheet10.append([codigo, None, None, None, titulo, precio, fecha, tipo])
        else:
            sheet10.append([codigo, None, None, None,titulo, precio, fecha, tipo])

        if precio:
            try:
                total_precio += float(precio)
            except ValueError:
                pass
        #print(total_precio)

    # Añadir la fila del total al final
    sheet10.append(["", "TOTAL CAPITULOS","","","",total_precio, "", "", ""])
    


    sheet11 = workbook.create_sheet("CONF_DECIMALES")
    sheet11.append(["CODIGO DECIMALES", "VALOR","DESCRIPCIÓN CODIGO"])


    # Diccionario para las descripciones de los decimales
    descripciones_decimales = {
        "DN": "Decimales del campo número de partes iguales de la hoja de mediciones. Por defecto 2 decimales.",
        "DD": "Decimales de dimensiones de las tres magnitudes de la hoja de mediciones. Por defecto 2 decimales.",
        "DS": "Decimales de la línea de subtotal o total de mediciones. Por defecto 2 decimales.",
        "DR": "Decimales de rendimiento y factor en una descomposición. Por defecto 3 decimales.",
        "DI": "Decimales del importe resultante de multiplicar rendimiento x precio del concepto. Por defecto 2 decimales.",
        "DP": "Decimales del importe resultante del sumatorio de los costes directos del concepto. Por defecto 2 decimales.",
        "DC": "Decimales del importe total del concepto. (CD+CI). Por defecto 2 decimales.",
        "DM": "Decimales del importe resultante de multiplicar la medición total del concepto por su precio. Por defecto 2 decimales.",
        "DIVISA": "Es la divisa expresada en el mismo modo que las abreviaturas utilizadas por el BCE (Banco Central Europeo), que en su caso deberán coincidir con las del registro ~V. En el Anexo 6 se indican las actuales.",
        "CI": "Costes Indirectos, expresados en porcentaje.",
        "GG": "Gastos Generales de la Empresa, expresados en porcentaje.",
        "BI": "Beneficio Industrial del contratista, expresado en porcentaje.",
        "BAJA": "Coeficiente de baja o alza de un presupuesto de adjudicación, expresado en porcentaje.",
        "IVA": "Impuesto del Valor Añadido, expresado en porcentaje.",
        "DRC": "Decimales del rendimiento y del factor de rendimiento de un presupuesto, y decimales del resultado de su multiplicación. Por defecto 3 decimales.",
        "DC2": "Decimales del importe de un presupuesto, de sus capitulos, subcapitulos, etc. y líneas de medición (unidades de obra excluidas), y decimales de los importes resultantes de multiplicar el rendimiento (o medición) total del presupuesto, sus capitulos, subcapitulos, etc. y líneas de medición (unidades de obra excluidas) por sus precios respectivos. Por defecto 2 decimales.",
        "DFS": "Decimales de los factores de rendimiento de las unidades de obra y de los elementos compuestos. Por defecto 3 decimales.",
        "DRS": "Decimales de los rendimientos de las unidades de obra y de los elementos compuestos, y decimales del resultado de la multiplicación de dichos rendimientos por sus respectivos factores. Por defecto 3 decimales.",
        "DUO": "Decimales del coste total de las unidades de obra. Por defecto 2 decimales.",
        "DI2": "Decimales de los importes resultantes de multiplicar los rendimientos totales de los elementos compuestos y/o elementos simples por sus respectivos precios, decimales del importe resultante del sumatorio de los costes directos de la unidad de obra y decimales de los costes indirectos. Decimales de los sumatorios sobre los que se aplican los porcentajes. Por defecto 2 decimales.",
        "DES": "Decimales del importe de los elementos simples. Por defecto 2 decimales.",
        "DN2": "Decimales del campo número de partes iguales de la hoja de mediciones. Por defecto 2 decimales.",
        "DD2": "Decimales de dimensiones de las tres magnitudes de la hoja de mediciones. Por defecto 2 decimales.",
        "DS2": "Decimales del total de mediciones. Por defecto 2 decimales.",
        "DSP": "Decimales de la línea de subtotal de mediciones. Por defecto 2 decimales.",
        "DEC": "Decimales del importe de los elementos compuestos. Por defecto 2.",
        "DIVISA2": "Es la divisa expresada en el mismo modo que las abreviaturas utilizadas por el BCE (Banco Central Europeo), que en su caso deberán coincidir con las del registro ~V. En el Anexo 6 se indican las actuales."
    }

    for registro_K, valores in registros_K.items():
        for encabezado, valor in valores.items():
            #if valor.strip():  # Verificar que el valor no esté vacío
            if isinstance(valor, str) and es_numero(valor):
                try:
                    valor = float(valor)
                except ValueError:
                    pass  # Si no se puede convertir a float, dejar como está

            # Obtener la descripción del encabezado
            descripcion = descripciones_decimales.get(encabezado, "")

            # Añadir la fila a la hoja de Excel
            sheet11.append([encabezado, valor, descripcion])



    base_nombre_archivo = os.path.basename(nombre_archivo)
    nombre_archivo_sin_extension = os.path.splitext(base_nombre_archivo)[0]

    directorio_salida = os.path.join(os.path.dirname(nombre_archivo), nombre_archivo_sin_extension)
    os.makedirs(directorio_salida, exist_ok=True)

    archivo_excel = os.path.join(directorio_salida, f"{nombre_archivo_sin_extension}.xlsx")
    workbook.save(archivo_excel)
    print(f"Archivo Excel guardado en: {archivo_excel}")

def seleccionar_archivo_bc3():
    root = tk.Tk()
    root.withdraw()
    ruta_archivo = filedialog.askopenfilename(
        title="Seleccionar archivo BC3",
        filetypes=[("Archivos BC3", "*.bc3")],
    )
    return ruta_archivo

def main():
    archivo_bc3 = seleccionar_archivo_bc3()
    if archivo_bc3:
        informacion_v = leer_primer_registro_V(archivo_bc3)
        contador_por_tipo, total_registros = contar_registros_por_tipo(archivo_bc3)
        textos_descriptivos = leer_registros_T(archivo_bc3)
        sin_clasificar_datos=leer_registros_C_sin_clasificar(archivo_bc3, textos_descriptivos)
        mano_obra_datos = leer_registros_C_mano_obra(archivo_bc3, textos_descriptivos)
        materiales_datos = leer_registros_C_materiales(archivo_bc3, textos_descriptivos)
        maquinaria_datos = leer_registros_C_Maquinaria(archivo_bc3, textos_descriptivos)
        residuos_datos=leer_registros_C_residuos(archivo_bc3, textos_descriptivos)
        capitulos_datos, nombre_medicion = extraer_lineas_capitulo(archivo_bc3)
        registros_K = extraer_registros_K(archivo_bc3)
        datos_C=leer_registros_C(archivo_bc3)
        registros_M,total_medicion,parcial_codigo,parcial_capitulo = leer_registros_M(archivo_bc3, datos_C)


        escribir_en_excel(archivo_bc3, contador_por_tipo, total_registros, informacion_v, sin_clasificar_datos, mano_obra_datos,materiales_datos, maquinaria_datos,residuos_datos, textos_descriptivos, capitulos_datos, nombre_medicion, registros_K, registros_M,total_medicion,parcial_codigo,parcial_capitulo)
    else:
        print("No se seleccionó ningún archivo.")

if __name__ == "__main__":
    main()
