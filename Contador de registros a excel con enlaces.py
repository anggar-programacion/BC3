import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from tkinter import Tk
from tkinter.filedialog import askdirectory

def contar_registros_por_tipo(archivo):
    tipos_validos = "VKCDYRTPLQJWGEOMNIABF"
    contador_registros = {tipo: 0 for tipo in tipos_validos}
    total_registros = 0

    try:
        with open(archivo, 'rb') as f:
            for linea in f:
                if len(linea) >= 2 and linea[0:1] == b'~':
                    tipo_registro = chr(linea[1])
                    if tipo_registro in tipos_validos:
                        contador_registros[tipo_registro] += 1
                        total_registros += 1
    except FileNotFoundError:
        print(f"Error: Archivo {archivo} no encontrado.")
    except Exception as e:
        print(f"Error al leer el archivo {archivo}: {e}")
    
    return contador_registros, total_registros

def procesar_directorio(directorio):
    try:
        archivos_bc3 = [f for f in os.listdir(directorio) if f.endswith('.bc3')]
    except FileNotFoundError:
        print(f"Error: Directorio {directorio} no encontrado.")
        return []
    except Exception as e:
        print(f"Error al acceder al directorio {directorio}: {e}")
        return []
    
    resumen = []
    
    for archivo in archivos_bc3:
        ruta_archivo = os.path.join(directorio, archivo)
        contador_registros, total_registros = contar_registros_por_tipo(ruta_archivo)
        resumen.append((archivo, ruta_archivo, contador_registros, total_registros))
    
    return resumen

def generar_excel(directorio, resumen):
    if not resumen:
        print("No hay datos para escribir en el archivo Excel.")
        return

    nombre_archivo_excel = os.path.basename(directorio.rstrip('/\\')) + '.xlsx'
    ruta_archivo_excel = os.path.join(directorio, nombre_archivo_excel)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    
    # Crear la cabecera
    header = ["Archivo"] + list(resumen[0][2].keys()) + ["Total"]
    ws.append(header)
    
    # Aplicar estilos a la cabecera
    font_bold = Font(bold=True)
    fill_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = font_bold
        cell.fill = fill_blue
    
    # Agregar los datos con hipervínculos
    for archivo, ruta_archivo, contador_registros, total_registros in resumen:
        fila = [archivo] + list(contador_registros.values()) + [total_registros]
        ws.append(fila)
        # Aplicar hipervínculo a la primera celda de la fila
        cell = ws.cell(row=ws.max_row, column=1)
        cell.hyperlink = ruta_archivo
        cell.style = 'Hyperlink'
    
    # Guardar el archivo
    try:
        wb.save(ruta_archivo_excel)
        print(f"Archivo Excel guardado en {ruta_archivo_excel}")
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")

def main():
    # Crear una ventana Tkinter para seleccionar el directorio
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal
    directorio = askdirectory(title="Selecciona el directorio con los archivos .bc3")
    
    if not directorio:
        print("No se seleccionó ningún directorio.")
        return

    resumen = procesar_directorio(directorio)
    generar_excel(directorio, resumen)

# Ejecutar el programa
if __name__ == "__main__":
    main()
